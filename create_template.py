#!/usr/bin/env python3
"""
Script để tạo file Excel template cho Google Sheets
File này có thể được import trực tiếp vào Google Sheets
"""

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

def create_excel_template():
    # Tạo workbook mới
    wb = openpyxl.Workbook()
    
    # Xóa sheet mặc định
    if 'Sheet' in wb.sheetnames:
        wb.remove(wb['Sheet'])
    
    # ========== SHEET 1: CONFIG ==========
    ws_config = wb.create_sheet("CONFIG")
    
    # Header
    ws_config['A1'] = "BHXH (%)"
    ws_config['B1'] = 0.08
    ws_config['A2'] = "BHYT (%)"
    ws_config['B2'] = 0.015
    ws_config['A3'] = "BHTN (%)"
    ws_config['B3'] = 0.01
    ws_config['A4'] = "Giảm trừ bản thân"
    ws_config['B4'] = 11000000
    
    # Format header
    header_fill = PatternFill(start_color="4CAF50", end_color="4CAF50", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")
    
    for row in ws_config.iter_rows(min_row=1, max_row=4, min_col=1, max_col=2):
        for cell in row:
            if cell.row == 1 or cell.column == 1:
                cell.font = header_font
                if cell.row == 1:
                    cell.fill = header_fill
    
    ws_config.column_dimensions['A'].width = 20
    ws_config.column_dimensions['B'].width = 15
    
    # ========== SHEET 2: BANG_THUE_TNCN ==========
    ws_tax = wb.create_sheet("BANG_THUE_TNCN")
    
    # Header
    headers_tax = ["Bậc", "Đến mức", "Thuế suất"]
    for col, header in enumerate(headers_tax, start=1):
        cell = ws_tax.cell(row=2, column=col)
        cell.value = header
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center')
    
    # Data
    tax_data = [
        [1, 5000000, "5%"],
        [2, 10000000, "10%"],
        [3, 18000000, "15%"],
        [4, 32000000, "20%"],
        [5, 52000000, "25%"],
        [6, 80000000, "30%"],
        [7, 999999999, "35%"]
    ]
    
    for row_idx, row_data in enumerate(tax_data, start=3):
        for col_idx, value in enumerate(row_data, start=1):
            cell = ws_tax.cell(row=row_idx, column=col_idx)
            cell.value = value
            if col_idx == 2:  # Format số tiền
                cell.number_format = '#,##0'
    
    ws_tax.column_dimensions['A'].width = 10
    ws_tax.column_dimensions['B'].width = 15
    ws_tax.column_dimensions['C'].width = 15
    
    # ========== SHEET 3: Danh_sach_nhan_vien ==========
    ws_employees = wb.create_sheet("Danh_sach_nhan_vien")
    
    # Header
    headers_emp = ["Mã NV", "Họ tên", "Email", "Lương GROSS", "Gửi template"]
    for col, header in enumerate(headers_emp, start=1):
        cell = ws_employees.cell(row=2, column=col)
        cell.value = header
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center')
    
    # Data
    employees_data = [
        ["NV01", "Nguyễn Văn A", "a@gmail.com", 20000000],
        ["NV02", "Trần Thị B", "b@gmail.com", 15000000],
        ["NV03", "Nguyễn Văn C", "c@gmail.com", 100000000],
        ["NV04", "Nguyễn Văn D", "d@gmail.com", 30000000]
    ]
    
    for row_idx, row_data in enumerate(employees_data, start=3):
        for col_idx, value in enumerate(row_data, start=1):
            cell = ws_employees.cell(row=row_idx, column=col_idx)
            cell.value = value
            if col_idx == 4:  # Format lương
                cell.number_format = '#,##0'
    
    # Cột "Gửi template" để trống (sẽ thêm checkbox sau khi import vào Google Sheets)
    
    ws_employees.column_dimensions['A'].width = 12
    ws_employees.column_dimensions['B'].width = 20
    ws_employees.column_dimensions['C'].width = 25
    ws_employees.column_dimensions['D'].width = 15
    ws_employees.column_dimensions['E'].width = 15
    
    # ========== SHEET 4: PHIEU_LUONG (Template) ==========
    ws_salary = wb.create_sheet("PHIEU_LUONG")
    
    # Tạo template phiếu lương
    salary_template = [
        ["A", "B", "C"],
        ["Mã NV", "", ""],
        ["Lương GROSS", "", ""],
        ["BHXH", "", ""],
        ["BHYT", "", ""],
        ["BHTN", "", ""],
        ["Thu nhập chịu thuế", "", ""],
        ["Thuế TNCN", "", ""],
        ["LƯƠNG NET", "", ""]
    ]
    
    for row_idx, row_data in enumerate(salary_template, start=1):
        for col_idx, value in enumerate(row_data, start=1):
            cell = ws_salary.cell(row=row_idx, column=col_idx)
            cell.value = value
            if row_idx == 1:  # Header row
                cell.font = header_font
                cell.fill = header_fill
            elif row_idx == 2 and col_idx == 1:  # Mã NV label
                cell.font = Font(bold=True)
            elif row_idx == 9 and col_idx == 1:  # LƯƠNG NET label
                cell.font = Font(bold=True)
                cell.fill = PatternFill(start_color="E8F5E9", end_color="E8F5E9", fill_type="solid")
    
    # Thêm công thức vào cột B (sẽ được Google Sheets convert)
    # Note: Excel và Google Sheets có syntax khác nhau, nên sẽ hướng dẫn thêm công thức sau khi import
    
    ws_salary.column_dimensions['A'].width = 20
    ws_salary.column_dimensions['B'].width = 20
    ws_salary.column_dimensions['C'].width = 10
    
    # ========== SHEET 5-8: NV01, NV02, NV03, NV04 ==========
    employees = ["NV01", "NV02", "NV03", "NV04"]
    employee_names = ["Nguyễn Văn A", "Trần Thị B", "Nguyễn Văn C", "Nguyễn Văn D"]
    gross_salaries = [20000000, 15000000, 100000000, 30000000]
    
    for idx, (emp_id, emp_name, gross) in enumerate(zip(employees, employee_names, gross_salaries)):
        ws_emp = wb.create_sheet(emp_id)
        
        # Header
        headers = ["A", "B", "C"]
        for col, header in enumerate(headers, start=1):
            cell = ws_emp.cell(row=1, column=col)
            cell.value = header
            cell.font = header_font
            cell.fill = header_fill
        
        # Labels và giá trị
        labels = [
            "Mã NV",
            "Lương GROSS",
            "BHXH",
            "BHYT",
            "BHTN",
            "Thu nhập chịu thuế",
            "Thuế TNCN",
            "LƯƠNG NET"
        ]
        
        for row_idx, label in enumerate(labels, start=2):
            cell_label = ws_emp.cell(row=row_idx, column=1)
            cell_label.value = label
            if row_idx == 2 or row_idx == 9:
                cell_label.font = Font(bold=True)
        
        # Giá trị Mã NV
        ws_emp['B2'] = emp_id
        
        # Giá trị Lương GROSS
        ws_emp['B3'] = gross
        ws_emp['B3'].number_format = '#,##0'
        
        # Các công thức sẽ được thêm sau khi import vào Google Sheets
        # Tạm thời để trống các cell B4-B9
        
        ws_emp.column_dimensions['A'].width = 20
        ws_emp.column_dimensions['B'].width = 20
        ws_emp.column_dimensions['C'].width = 10
        
        # Highlight row LƯƠNG NET
        for col in range(1, 4):
            cell = ws_emp.cell(row=9, column=col)
            if col == 1:
                cell.fill = PatternFill(start_color="E8F5E9", end_color="E8F5E9", fill_type="solid")
    
    # Lưu file
    filename = "Template_Quan_Ly_Luong.xlsx"
    wb.save(filename)
    print(f"✅ Đã tạo file: {filename}")
    print(f"\n📋 File bao gồm các sheet:")
    print("   1. CONFIG - Cấu hình chung")
    print("   2. BANG_THUE_TNCN - Bảng thuế TNCN")
    print("   3. Danh_sach_nhan_vien - Danh sách nhân viên")
    print("   4. PHIEU_LUONG - Template phiếu lương")
    print("   5-8. NV01, NV02, NV03, NV04 - Phiếu lương từng nhân viên")
    print(f"\n📤 Bước tiếp theo:")
    print("   1. Mở Google Sheets")
    print("   2. File > Import > Upload")
    print("   3. Chọn file {filename}")
    print("   4. Chọn 'Replace spreadsheet' hoặc 'Insert new sheet(s)'")
    print("   5. Thêm công thức vào các sheet NV01-NV04 (xem hướng dẫn trong file HUONG_DAN_FORMULA.md)")

if __name__ == "__main__":
    try:
        create_excel_template()
    except ImportError:
        print("❌ Lỗi: Cần cài đặt thư viện openpyxl")
        print("   Chạy lệnh: pip install openpyxl")
    except Exception as e:
        print(f"❌ Lỗi: {e}")
